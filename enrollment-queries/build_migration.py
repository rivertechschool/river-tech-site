"""Build migration rows from the Cognito Homeschool export."""
import openpyxl
import json
from datetime import datetime

XLSX = '/sessions/trusting-zealous-gates/mnt/danhegelund/Documents/Claude Cowork/HomeschoolEnrollSeptember2ndStart.xlsx'
wb = openpyxl.load_workbook(XLSX, data_only=True)

main_ws = wb['HomeschoolEnrollSeptember2ndSta']
main_headers = [c.value for c in main_ws[1]]
main_rows = {}
for row_idx in range(2, main_ws.max_row + 1):
    row = {h: main_ws.cell(row_idx, i).value for i, h in enumerate(main_headers, 1)}
    main_rows[str(row['HomeschoolEnrollSeptember2ndSta_Id'])] = row

fam_ws = wb['FamilyInformation']
fam_headers = [c.value for c in fam_ws[1]]
families = {}
for row_idx in range(2, fam_ws.max_row + 1):
    row = {h: fam_ws.cell(row_idx, i).value for i, h in enumerate(fam_headers, 1)}
    eid = str(row['HomeschoolEnrollSeptember2ndSta_Id'])
    families.setdefault(eid, []).append(row)

child_ws = wb['ChildInformationAndDetails']
child_headers = [c.value for c in child_ws[1]]
kids = {}
for row_idx in range(2, child_ws.max_row + 1):
    row = {h: child_ws.cell(row_idx, i).value for i, h in enumerate(child_headers, 1)}
    eid = str(row['HomeschoolEnrollSeptember2ndSta_Id'])
    kids.setdefault(eid, []).append(row)

GRADE_MAP = {'2-4': 'elementary', '5-6': 'elementary', '7-8': 'middle', 'High School': 'high'}
PROG_MAP = {
    'Monday: Performing Arts': 'Monday',
    'Tuesday: Science & Social Studies': 'Tuesday',
    'Thursday: Life Skills': 'Thursday',
    'Friday: Technology': 'Friday',
}
PREV_TOKENS = {
    'Public School': 'public',
    'Homeschool': 'homeschool',
    '"Another Private School (If so, which)"': 'private',
    'Another Private School (If so, which)': 'private',
    'Other (please describe)': 'other',
}

def fmt_dt(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    return v or ''

def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return v or ''

def fmt_address(f):
    line1 = f.get('HomeAddress_Line1') or ''
    if f.get('HomeAddress_Line2'):
        line1 += ', ' + f['HomeAddress_Line2']
    csz = ', '.join(p for p in [f.get('HomeAddress_City') or '', f.get('HomeAddress_State') or '', f.get('HomeAddress_PostalCode') or ''] if p)
    return ((line1 + ', ') if line1 else '') + csz

def map_programs(s):
    if not s:
        return ''
    out = []
    for chunk in [c.strip() for c in s.split(',')]:
        for cog, our in PROG_MAP.items():
            if chunk == cog:
                out.append(our)
                break
    return ', '.join(out)

def map_prev(s):
    if not s:
        return ('', '')
    parts = []
    buf = ''
    in_quote = False
    for ch in s:
        if ch == '"':
            in_quote = not in_quote
            buf += ch
        elif ch == ',' and not in_quote:
            parts.append(buf.strip())
            buf = ''
        else:
            buf += ch
    if buf.strip():
        parts.append(buf.strip())
    tokens = []
    for p in parts:
        if p in PREV_TOKENS:
            tokens.append(PREV_TOKENS[p])
    return (', '.join(tokens), '')

def max_days_int(v):
    if not v:
        return ''
    try:
        return int(str(v).split()[0])
    except Exception:
        return v

output = []
for eid, mrow in sorted(main_rows.items(), key=lambda x: int(x[0])):
    entry_status = mrow.get('Entry_Status', '')
    payment_status = mrow.get('Order_PaymentStatus', '')
    if entry_status == 'Incomplete':
        continue
    fams = families.get(eid, [])
    children = kids.get(eid, [])
    # Sort children by DOB (oldest first) — deterministic ordering
    children.sort(key=lambda c: (c.get('DateOfBirth') or datetime.max, c.get('Name_First') or ''))
    p1 = fams[0] if len(fams) >= 1 else {}
    p2 = fams[1] if len(fams) >= 2 else {}

    parent1_name = ((p1.get('ParentGuardianFullName_First') or '') + ' ' + (p1.get('ParentGuardianFullName_Last') or '')).strip()

    row = {
        'Registration ID': f'MIG-{eid}',
        'Submitted (UTC)': fmt_dt(mrow.get('Entry_DateSubmitted')),
        'Status': payment_status or 'Migrated',
        'Parent 1 First': p1.get('ParentGuardianFullName_First') or '',
        'Parent 1 Last': p1.get('ParentGuardianFullName_Last') or '',
        'Parent 1 Email': p1.get('EmailAddress') or '',
        'Parent 1 Phone': p1.get('ContactPhoneNumber') or '',
        'Parent 1 Address': fmt_address(p1),
        'Parent 2 First': p2.get('ParentGuardianFullName_First') or '',
        'Parent 2 Last': p2.get('ParentGuardianFullName_Last') or '',
        'Parent 2 Email': p2.get('EmailAddress') or '',
        'Parent 2 Phone': p2.get('ContactPhoneNumber') or '',
        'Children Count': mrow.get('StudentEnrollmentDetails_NumberOfChildrenBeingEnrolled') or '',
        'Max Days': max_days_int(mrow.get('HowManyDaysIsYourHighestenrolledChildAttending')),
        'Family Fee (USD)': mrow.get('AnnualFamilySetupFee') or '',
        'Signature': parent1_name,
        'Signature Date': fmt_date(mrow.get('SignatureAndSubmission_TodaysDate')),
    }
    for i in range(1, 7):
        k = children[i-1] if len(children) >= i else None
        if k:
            grade_raw = k.get('YourChildsGradeLevel') or ''
            grade = GRADE_MAP.get(grade_raw, '')
            prev_list, prev_other = map_prev(k.get('WhatIsYourChildsPreviousSchooling'))
            row[f'C{i} Name'] = ((k.get('Name_First') or '') + ' ' + (k.get('Name_Last') or '')).strip()
            row[f'C{i} DOB'] = fmt_date(k.get('DateOfBirth'))
            row[f'C{i} Gender'] = k.get('Gender') or ''
            row[f'C{i} Grade'] = grade
            row[f'C{i} Reading'] = 'Unknown - migrated' if grade == 'elementary' else ''
            row[f'C{i} Tablet'] = 'Unknown - migrated' if grade == 'elementary' else ''
            row[f'C{i} Programs'] = map_programs(k.get('PROGRAMSSELECTED'))
            row[f'C{i} Previous Schooling'] = prev_list
            row[f'C{i} Prev Schooling Other'] = k.get('SpaceToDescribeIfNeeded') or ''
            row[f'C{i} Attitude'] = k.get('DescribeYourChildsAttitudeTowardsLearningAndTheirAbilityToWorkIndependently') or ''
            row[f'C{i} Health'] = k.get('IfApplicablePleaseDescribeAnyHealthConcernsOrSpecialNeedsYourChildHas') or ''
            row[f'C{i} Hopes'] = k.get('WhatDoYouHopeYourChildWillGainFromTheirExperienceAtRiverTech') or ''
            row[f'C{i} Notes'] = k.get('AdditionalInformationAboutChild1') or ''
            row[f'C{i} Photo URL'] = ''
        else:
            for fld in ['Name', 'DOB', 'Gender', 'Grade', 'Reading', 'Tablet', 'Programs', 'Previous Schooling', 'Prev Schooling Other', 'Attitude', 'Health', 'Hopes', 'Notes', 'Photo URL']:
                row[f'C{i} {fld}'] = ''
    row['Cognito Entry ID'] = eid
    row['Cognito Order ID'] = mrow.get('Order_Id') or ''
    row['Stripe Charge ID'] = mrow.get('Order_PaymentConfirmationNumber') or ''
    row['Amount Paid'] = mrow.get('Order_AmountPaid') or ''
    row['Payment Date'] = fmt_dt(mrow.get('Order_PaymentDate'))
    row['Payment Method'] = mrow.get('Order_PaymentMethod') or ''
    row['Migrated'] = 'Yes'
    output.append(row)

with open('/sessions/trusting-zealous-gates/migration/migration_rows.json', 'w') as f:
    json.dump(output, f, indent=2, default=str)

print(f'Built {len(output)} migration rows (skipped Incomplete entries)')
print()
print('=== SAMPLE ROW 1 (Cognito entry 20 - Lovoi family, 1 kid) ===')
r = output[-1]  # last built is entry 20 since we sort ascending
for k, v in r.items():
    if v and not any(k.startswith(f'C{i} ') for i in range(1, 7)):
        print(f'  {k}: {repr(v)[:140]}')
print('  --- child 1 ---')
for k, v in r.items():
    if k.startswith('C1 ') and v:
        print(f'    {k}: {repr(v)[:200]}')

print()
print('=== SAMPLE ROW 2 (Cognito entry 19 - Stivers family, 3 kids, 2 parents) ===')
r_stivers = next(x for x in output if x['Cognito Entry ID'] == '19')
for k, v in r_stivers.items():
    if v and not any(k.startswith(f'C{i} ') for i in range(1, 7)):
        print(f'  {k}: {repr(v)[:140]}')
for ci in range(1, 4):
    print(f'  --- child {ci} ---')
    for k, v in r_stivers.items():
        if k.startswith(f'C{ci} ') and v:
            print(f'    {k}: {repr(v)[:200]}')
